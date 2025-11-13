"""
Модуль для обработки Excel файлов
"""
import pandas as pd
import logging
from typing import List, Dict, Any, Tuple
from decimal import Decimal, InvalidOperation
import io
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelProcessor:
    def __init__(self):
        """Инициализация процессора Excel"""
        self.supported_formats = ['.xlsx', '.xls', '.xlsm', '.csv']
        self.income_categories = [
            "Входные билеты",
            "Бар",
            "Консумация Бара",
            "Консумация кухни",
            "Crazy Menu",
            "Общий чай",
            "Overtime",
            "Кальяны",
            "Шары",
            "Штрафы",
            "Стафф",
            "Стафф кальян",
            "Доход клуба",
            "Сервисный сбор",
            "Итого",
            "плюс по кассе",
            "Итого за смену"
        ]
    
    def process_file(self, file_content: bytes, file_name: str) -> Tuple[List[Dict[str, Any]], str]:
        """
        Обработка Excel файла
        
        Returns:
            Tuple[List[Dict], str]: (данные в виде списка словарей, краткая статистика)
        """
        try:
            # Определение типа файла
            if file_name.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(file_content))
            else:
                df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
            
            # Очистка данных
            df = self._clean_dataframe(df)
            
            # Преобразование в список словарей
            data = df.to_dict('records')
            
            # Генерация статистики
            stats = self._generate_statistics(df)
            
            logger.info(f"Processed file {file_name}: {len(data)} rows, {len(df.columns)} columns")
            
            return data, stats
        
        except Exception as e:
            logger.error(f"Error processing file {file_name}: {e}")
            raise ValueError(f"Не удалось обработать файл: {str(e)}")

    @staticmethod
    def _parse_decimal(value) -> Decimal:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return Decimal('0')

        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value)).quantize(Decimal('0.01'))

        if isinstance(value, str):
            cleaned = re.sub(r'[^0-9,\.\-]', '', value)
            cleaned = cleaned.replace(' ', '').replace(',', '.').strip()
            if cleaned == '':
                return Decimal('0')
            try:
                return Decimal(cleaned).quantize(Decimal('0.01'))
            except InvalidOperation:
                logger.warning(f"Failed to parse decimal from string '{value}'")
                return Decimal('0')

        logger.warning(f"Unsupported value type for decimal parsing: {value} ({type(value)})")
        return Decimal('0')

    def extract_income_records(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Извлечение блока «Доходы» с первого листа"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=None, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading Excel for income block: {e}")
            return []

        if df.empty:
            return []

        # Ищем ДОХОДЫ в первой строке (горизонтальный формат с несколькими блоками)
        income_col = None
        if df.shape[0] > 0:
            for col_idx in range(df.shape[1]):
                cell_value = df.iloc[0, col_idx]
                if isinstance(cell_value, str) and 'ДОХОД' in cell_value.strip().upper():
                    income_col = col_idx
                    logger.info(f"Found 'ДОХОДЫ' header in column {col_idx}")
                    break
        
        # Если найден горизонтальный формат
        if income_col is not None:
            return self._extract_income_horizontal(df, income_col)
        
        # Иначе ищем вертикальный формат (старая логика)
        start_row = None
        for idx, value in enumerate(df.iloc[:, 0]):
            if isinstance(value, str) and value.strip().upper() == 'ДОХОДЫ':
                start_row = idx + 1
                logger.info(f"Found 'ДОХОДЫ' header at row {idx}, data starts at row {start_row}")
                break

        if start_row is None:
            logger.info("Income block header 'ДОХОДЫ' not found")
            return []
        
        return self._extract_income_vertical(df, start_row)
    
    def _extract_income_horizontal(self, df: pd.DataFrame, income_col: int) -> List[Dict[str, Any]]:
        """Извлечение доходов из горизонтального формата - универсальный подход"""
        records: List[Dict[str, Any]] = []
        
        # Универсальная логика: в каждой строке ищем текст (категория), потом первое число справа (сумма)
        for row_idx in range(1, len(df)):
            # Читаем категорию из колонки income_col
            raw_category = df.iloc[row_idx, income_col] if df.shape[1] > income_col else None
            
            if raw_category is None or (isinstance(raw_category, float) and pd.isna(raw_category)):
                # Пустая строка - конец блока
                break
            
            # ВАЖНО: Если в колонке категории ЧИСЛО - это не доход, а другой блок (например, цены билетов)
            if isinstance(raw_category, (int, float)):
                logger.info(f"Row {row_idx}: col {income_col} contains number {raw_category}, stopping income parsing")
                break
            
            category = str(raw_category).strip()
            if not category:
                break
            
            # Останавливаемся на "Итого за смену"
            category_upper = category.upper()
            if 'ИТОГО' in category_upper and 'СМЕН' in category_upper:
                # Ищем первое число справа от категории
                amount = None
                for col_offset in range(1, 6):
                    if df.shape[1] > income_col + col_offset:
                        candidate = df.iloc[row_idx, income_col + col_offset]
                        if candidate is not None and not (isinstance(candidate, float) and pd.isna(candidate)):
                            # Проверяем, что это число, а не текст
                            if isinstance(candidate, (int, float)) or (isinstance(candidate, str) and candidate.replace('.', '').replace(',', '').replace('-', '').isdigit()):
                                amount = self._parse_decimal(candidate)
                                logger.info(f"Income ИТОГО: category='{category}', amount={amount} (col offset {col_offset})")
                                break
                
                if amount is not None:
                    records.append({'category': category, 'amount': amount})
                break
            
            # Ищем первое ЧИСЛО справа от категории (пропускаем пустые ячейки)
            amount = None
            for col_offset in range(1, 6):
                if df.shape[1] > income_col + col_offset:
                    candidate = df.iloc[row_idx, income_col + col_offset]
                    if candidate is not None and not (isinstance(candidate, float) and pd.isna(candidate)):
                        # Проверяем, что это число
                        if isinstance(candidate, (int, float)):
                            amount = self._parse_decimal(candidate)
                            logger.info(f"Income: category='{category}', amount={amount} (found at col offset {col_offset})")
                            break
                        elif isinstance(candidate, str):
                            # Проверяем, не начинается ли следующий блок (текст вместо числа)
                            if not candidate.replace('.', '').replace(',', '').replace('-', '').replace(' ', '').isdigit():
                                logger.info(f"Stopped at '{category}' - next column contains text '{candidate}'")
                                break
                            else:
                                amount = self._parse_decimal(candidate)
                                logger.info(f"Income: category='{category}', amount={amount}")
                                break
            
            # Если НЕ нашли число справа - проверяем, не начался ли другой блок
            if amount is None:
                # Специально для "Входные билеты" без суммы - это начало отдельного блока
                if category_upper == 'ВХОДНЫЕ БИЛЕТЫ':
                    logger.info(f"Found 'Входные билеты' without amount at row {row_idx}, stopping income parsing")
                    break
                # Для других категорий без суммы - сохраняем с нулевой суммой!
                logger.info(f"No amount found for '{category}', saving with 0")
                amount = Decimal('0')
            
            records.append({
                'category': category,
                'amount': amount
            })
        
        return records
    
    def _extract_income_vertical(self, df: pd.DataFrame, start_row: int) -> List[Dict[str, Any]]:
        """Извлечение доходов из вертикального формата (заголовок в первой колонке)"""
        records: List[Dict[str, Any]] = []
        observed_categories = set()

        for row_idx in range(start_row, len(df)):
            raw_category = df.iloc[row_idx, 0] if df.shape[1] > 0 else None
            raw_amount = df.iloc[row_idx, 1] if df.shape[1] > 1 else None

            if raw_category is None or (isinstance(raw_category, float) and pd.isna(raw_category)):
                break

            category = str(raw_category).strip()

            if category.upper() not in {name.upper() for name in self.income_categories}:
                break

            amount = self._parse_decimal(raw_amount)
            logger.info(f"Income (vertical): category='{category}', raw_amount={raw_amount}, parsed={amount}")
            
            records.append({
                'category': category,
                'amount': amount
            })
            observed_categories.add(category.upper())

        return records
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Очистка DataFrame от пустых строк и нормализация имен колонок"""
        # Удаление пустых строк
        df = df.dropna(how='all')
        
        # Нормализация имен колонок
        df.columns = [str(col).strip().replace(' ', '_').lower() for col in df.columns]
        
        # Замена NaN на None для совместимости с SQL
        df = df.where(pd.notna(df), None)
        
        return df
    
    def _generate_statistics(self, df: pd.DataFrame) -> str:
        """Генерация статистики по данным"""
        stats_lines = [
            f"📊 **Статистика файла:**",
            f"",
            f"🔢 Количество строк: {len(df)}",
            f"📝 Количество колонок: {len(df.columns)}",
            f"",
            f"**Колонки:**"
        ]
        
        for col in df.columns:
            # Подсчет непустых значений
            non_null_count = df[col].notna().sum()
            
            # Определение типа данных
            if pd.api.types.is_numeric_dtype(df[col]):
                dtype = "Числовой"
                # Статистика для числовых данных
                try:
                    min_val = df[col].min()
                    max_val = df[col].max()
                    avg_val = df[col].mean()
                    stats_lines.append(
                        f"  • **{col}** ({dtype}): {non_null_count} значений | "
                        f"Мин: {min_val:.2f}, Макс: {max_val:.2f}, Среднее: {avg_val:.2f}"
                    )
                except:
                    stats_lines.append(f"  • **{col}** ({dtype}): {non_null_count} значений")
            else:
                dtype = "Текстовый"
                unique_count = df[col].nunique()
                stats_lines.append(
                    f"  • **{col}** ({dtype}): {non_null_count} значений | "
                    f"Уникальных: {unique_count}"
                )
        
        return "\n".join(stats_lines)
    
    def get_column_info(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Получение информации о колонках для AI"""
        if not data:
            return {}
        
        columns_info = {}
        sample_data = data[:5]  # Берем первые 5 строк как образец
        
        for col in data[0].keys():
            sample_values = [row[col] for row in sample_data if row.get(col) is not None]
            columns_info[col] = {
                'sample_values': sample_values[:3],  # Первые 3 значения
                'type': type(sample_values[0]).__name__ if sample_values else 'unknown'
            }
        
        return columns_info
    
    def validate_file(self, file_name: str) -> bool:
        """Проверка поддерживаемого формата файла"""
        return any(file_name.lower().endswith(fmt) for fmt in self.supported_formats)
    
    def export_to_excel(self, data: List[Dict[str, Any]], file_name: str = "export.xlsx") -> bytes:
        """Экспорт данных обратно в Excel"""
        try:
            df = pd.DataFrame(data)
            
            # Создание Excel файла в памяти
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data')
            
            output.seek(0)
            return output.getvalue()
        
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise ValueError(f"Не удалось экспортировать данные: {str(e)}")
    
    def export_full_period_report_to_excel(self, all_blocks: Dict[str, List[Dict[str, Any]]], club_name: str, start_date, end_date) -> bytes:
        """Экспорт ПОЛНОГО комплексного отчета за период в Excel со всеми блоками"""
        try:
            from datetime import date
            from openpyxl import Workbook
            from openpyxl.styles import Font
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Полный отчет"
            
            # Заголовок отчета
            ws['A1'] = f'Клуб: {club_name}'
            ws['A1'].font = Font(bold=True, size=14)
            
            start_str = start_date.strftime("%d.%m.%Y") if isinstance(start_date, date) else str(start_date)
            end_str = end_date.strftime("%d.%m.%Y") if isinstance(end_date, date) else str(end_date)
            ws['B1'] = f'Период: {start_str} - {end_str}'
            ws['B1'].font = Font(bold=True, size=14)
            
            current_row = 3  # Начинаем с 3-й строки
            
            bold_font = Font(bold=True, size=11)
            block_title_font = Font(bold=True, size=13)
            
            # Обрабатываем каждый блок
            for block_name, block_data in all_blocks.items():
                if not block_data:
                    continue
                
                # Заголовок блока
                ws.cell(row=current_row, column=1, value=f"📊 {block_name.upper()}")
                ws.cell(row=current_row, column=1).font = block_title_font
                current_row += 1
                
                # Заголовки колонок
                if block_data:
                    headers = list(block_data[0].keys())
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=header)
                        cell.font = bold_font
                    current_row += 1
                    
                    # Данные блока
                    for row_data in block_data:
                        for col_idx, header in enumerate(headers, start=1):
                            value = row_data.get(header)
                            cell = ws.cell(row=current_row, column=col_idx, value=value)
                            
                            # Делаем строки с "ИТОГО" жирными
                            first_col_value = row_data.get(headers[0])
                            if first_col_value and isinstance(first_col_value, str) and 'итого' in first_col_value.lower():
                                cell.font = bold_font
                        
                        current_row += 1
                
                # Пустая строка между блоками
                current_row += 1
            
            # Сохраняем в память
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        
        except Exception as e:
            logger.error(f"Error exporting full period report to Excel: {e}")
            raise ValueError(f"Не удалось экспортировать комплексный отчет: {str(e)}")
    
    def export_period_report_to_excel(self, data: List[Dict[str, Any]], club_name: str, start_date, end_date, block_name: str) -> bytes:
        """Экспорт сводного отчета за период в Excel"""
        try:
            from datetime import date
            from openpyxl.styles import Font
            
            df = pd.DataFrame(data)
            
            # Создание Excel файла в памяти
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Записываем данные, начиная со строки 3 (оставляем место для заголовка)
                df.to_excel(writer, index=False, sheet_name='Data', startrow=2)
                
                # Получаем worksheet для добавления заголовка
                worksheet = writer.sheets['Data']
                
                # Добавляем заголовок
                worksheet['A1'] = f'Клуб: {club_name}'
                worksheet['A1'].font = Font(bold=True, size=12)
                
                # Форматируем период
                start_str = start_date.strftime("%d.%m.%Y") if isinstance(start_date, date) else str(start_date)
                end_str = end_date.strftime("%d.%m.%Y") if isinstance(end_date, date) else str(end_date)
                worksheet['B1'] = f'Период: {start_str} - {end_str}'
                worksheet['B1'].font = Font(bold=True, size=12)
                
                # Делаем строки с "итого" жирными (и текст, и цифры)
                bold_font = Font(bold=True, size=11)
                for row_idx in range(3, worksheet.max_row + 1):  # Начинаем с 3-й строки (данные)
                    cell_value = worksheet.cell(row=row_idx, column=1).value  # Колонка "Категория"
                    if cell_value and isinstance(cell_value, str) and 'итого' in cell_value.lower():
                        # Делаем жирным всю строку (категория + сумма)
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_idx, column=col_idx).font = bold_font
            
            output.seek(0)
            return output.getvalue()
        
        except Exception as e:
            logger.error(f"Error exporting period report to Excel: {e}")
            raise ValueError(f"Не удалось экспортировать данные: {str(e)}")

    def export_to_excel_with_header(self, data: List[Dict[str, Any]], report_date, block_name: str, club_name: str = None) -> bytes:
        """Экспорт данных в Excel с заголовком (дата, клуб и название блока)"""
        try:
            from datetime import date
            from openpyxl.styles import Font
            
            df = pd.DataFrame(data)
            
            # Создание Excel файла в памяти
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Записываем данные, начиная со строки 3 (оставляем место для заголовка)
                df.to_excel(writer, index=False, sheet_name='Data', startrow=2)
                
                # Получаем worksheet для добавления заголовка
                worksheet = writer.sheets['Data']
                
                # Добавляем заголовок в первую строку: Дата в A1, Клуб в B1
                date_str = report_date.strftime("%d.%m.%Y") if isinstance(report_date, date) else str(report_date)
                worksheet['A1'] = f'Дата: {date_str}'
                worksheet['A1'].font = Font(bold=True, size=12)
                
                if club_name:
                    worksheet['B1'] = f'Клуб: {club_name}'
                    worksheet['B1'].font = Font(bold=True, size=12)
            
            output.seek(0)
            return output.getvalue()
        
        except Exception as e:
            logger.error(f"Error exporting to Excel with header: {e}")
            raise ValueError(f"Не удалось экспортировать данные: {str(e)}")

    @staticmethod
    def _parse_int(value) -> int:
        decimal_value = ExcelProcessor._parse_decimal(value)
        return int(decimal_value)

    def extract_ticket_sales(self, file_content: bytes) -> Dict[str, Any]:
        """Извлечение блока «Входные билеты» с первого листа"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=None, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading Excel for ticket sales block: {e}")
            return {}

        if df.empty:
            return {}

        # Ищем блок "Входные билеты" - это отдельный блок, не часть доходов
        # Признак: следующая строка содержит заголовки "цена", "кол-во", "сумма"
        start_row = None
        for idx, value in enumerate(df.iloc[:, 0]):
            if isinstance(value, str) and value.strip().upper() == 'ВХОДНЫЕ БИЛЕТЫ':
                # Проверяем следующую строку - должна быть "цена | кол-во | сумма"
                if idx + 1 < len(df):
                    next_row_cells = [df.iloc[idx+1, col] if df.shape[1] > col else None for col in range(3)]
                    next_row_text = ' '.join([str(c).lower() for c in next_row_cells if pd.notna(c)])
                    if 'цена' in next_row_text and 'кол' in next_row_text:
                        start_row = idx + 1
                        logger.info(f"Found ticket sales block at row {idx}")
                        break

        if start_row is None:
            logger.info("Ticket sales block header 'ВХОДНЫЕ БИЛЕТЫ' not found")
            return {}

        header_row = None
        header_keywords = {'цена', 'кол', 'кол-во', 'количество', 'сумма'}

        for row_idx in range(start_row, len(df)):
            cells = [df.iloc[row_idx, col] if df.shape[1] > col else None for col in range(3)]
            normalized = [str(cell).strip().lower() if cell is not None and not (isinstance(cell, float) and pd.isna(cell)) else '' for cell in cells]

            if any('цена' in cell for cell in normalized) and any('кол' in cell for cell in normalized) and any('сумма' in cell for cell in normalized):
                header_row = row_idx
                start_row = row_idx + 1
                break

            if normalized[0]:
                # Если первая значимая строка после заголовка — данные, а не шапка
                header_row = None
                start_row = row_idx
                break

        records: List[Dict[str, Any]] = []
        calculated_quantity = 0
        calculated_amount = Decimal('0.00')
        reported_total_quantity = None
        reported_total_amount = None

        for row_idx in range(start_row, len(df)):
            price_cell = df.iloc[row_idx, 0] if df.shape[1] > 0 else None
            quantity_cell = df.iloc[row_idx, 1] if df.shape[1] > 1 else None
            amount_cell = df.iloc[row_idx, 2] if df.shape[1] > 2 else None

            # Проверяем на итоговую строку
            if price_cell is not None and isinstance(price_cell, str):
                normalized = price_cell.strip().upper()
                if 'ИТОГО' in normalized:
                    # Это итоговая строка
                    quantity = self._parse_int(quantity_cell)
                    amount = self._parse_decimal(amount_cell)
                    
                    record = {
                        'price_label': price_cell.strip(),
                        'price_value': None,
                        'quantity': quantity,
                        'amount': amount,
                        'is_total': True
                    }
                    records.append(record)
                    reported_total_quantity = quantity
                    reported_total_amount = amount
                    logger.info(f"Found ИТОГО: qty={quantity}, amt={amount}")
                    break

            # Пропускаем пустые строки (продолжаем искать ИТОГО)
            if price_cell is None or (isinstance(price_cell, float) and pd.isna(price_cell)):
                # Проверяем следующие несколько строк на наличие ИТОГО
                found_total = False
                for next_idx in range(row_idx + 1, min(row_idx + 5, len(df))):
                    next_cell = df.iloc[next_idx, 0]
                    if next_cell is not None and isinstance(next_cell, str) and 'ИТОГО' in next_cell.upper():
                        found_total = True
                        break
                if not found_total:
                    break
                else:
                    continue  # Пропускаем пустую строку, продолжаем поиск

            price_label = str(price_cell).strip()
            if not price_label:
                continue

            # Обычная строка с данными билетов
            parsed_price = self._parse_decimal(price_cell)
            quantity = self._parse_int(quantity_cell)
            amount = self._parse_decimal(amount_cell)

            record = {
                'price_label': price_label,
                'price_value': parsed_price,
                'quantity': quantity,
                'amount': amount,
                'is_total': False
            }

            records.append(record)
            calculated_quantity += quantity
            calculated_amount += amount

        if not records:
            return {}

        total_quantity = reported_total_quantity if reported_total_quantity is not None else calculated_quantity
        total_amount = reported_total_amount if reported_total_amount is not None else calculated_amount

        totals_match = True
        if reported_total_quantity is not None and reported_total_quantity != calculated_quantity:
            totals_match = False
        if reported_total_amount is not None and (reported_total_amount - calculated_amount).copy_abs() > Decimal('0.01'):
            totals_match = False

        return {
            'records': records,
            'calculated_quantity': calculated_quantity,
            'calculated_amount': calculated_amount,
            'total_quantity': total_quantity,
            'total_amount': total_amount,
            'totals_match': totals_match
        }

    def extract_payment_types(self, file_content: bytes) -> Dict[str, Any]:
        """Извлечение блока «Типы оплат за смену»"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=None, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading Excel for payment types block: {e}")
            return {}

        if df.empty:
            return {}

        start_row = None
        for idx, value in enumerate(df.iloc[:, 0]):
            if isinstance(value, str) and value.strip().upper() == 'НАЛИЧНЫЕ':
                start_row = idx
                break

        if start_row is None:
            logger.info("Payment types block header (cash) not found")
            return {}

        records: List[Dict[str, Any]] = []
        calculated_total = Decimal('0.00')
        reported_total = None
        reported_cash_total = None

        for row_idx in range(start_row, len(df)):
            label_cell = df.iloc[row_idx, 0] if df.shape[1] > 0 else None
            amount_cell = df.iloc[row_idx, 2] if df.shape[1] > 2 else None  # Колонка 2, не 1!

            if label_cell is None or (isinstance(label_cell, float) and pd.isna(label_cell)):
                # Пустая строка - проверяем, есть ли дальше ИТОГО
                found_total = False
                for next_idx in range(row_idx + 1, min(row_idx + 5, len(df))):
                    next_cell = df.iloc[next_idx, 0]
                    if next_cell is not None and isinstance(next_cell, str) and 'ИТОГО' in next_cell.upper():
                        found_total = True
                        break
                if not found_total:
                    break
                else:
                    continue

            label = str(label_cell).strip()
            if not label:
                continue

            normalized = label.upper()

            if normalized.startswith('ИТОГО КАССА'):
                reported_cash_total = self._parse_decimal(amount_cell)
                records.append({
                    'payment_type': 'ИТОГО КАССА',
                    'amount': reported_cash_total,
                    'is_total': False,
                    'is_cash_total': True
                })
                logger.info(f"Found ИТОГО КАССА: {reported_cash_total}")
                continue

            if normalized.startswith('ИТОГО'):
                reported_total = self._parse_decimal(amount_cell)
                records.append({
                    'payment_type': 'ИТОГО',
                    'amount': reported_total,
                    'is_total': True,
                    'is_cash_total': False
                })
                logger.info(f"Found ИТОГО: {reported_total}")
                break

            amount = self._parse_decimal(amount_cell)
            calculated_total += amount
            records.append({
                'payment_type': label,
                'amount': amount,
                'is_total': False,
                'is_cash_total': False
            })
            logger.info(f"Payment type: {label} = {amount}")

        if not records:
            return {}

        totals_match = True
        if reported_total is not None:
            if (reported_total - calculated_total).copy_abs() > Decimal('0.01'):
                totals_match = False
        else:
            reported_total = calculated_total

        return {
            'records': records,
            'calculated_total': calculated_total,
            'reported_total': reported_total,
            'cash_total': reported_cash_total,
            'totals_match': totals_match
        }

    def extract_staff_statistics(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Извлечение блока «Статистика персонала» - горизонтальный формат"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=None, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading Excel for staff statistics block: {e}")
            return []

        if df.empty:
            return []

        # Ищем заголовок блока
        start_row = None
        for idx, value in enumerate(df.iloc[:, 0]):
            if isinstance(value, str) and 'статистика' in value.strip().lower() and 'персонал' in value.strip().lower():
                start_row = idx + 1
                logger.info(f"Found 'Статистика персонала' at row {idx}, data starts at {start_row}")
                break

        if start_row is None:
            logger.info("Staff statistics block header not found")
            return []

        records: List[Dict[str, Any]] = []

        # Данные идут горизонтально: колонка 0 - должность, колонка 2 - количество
        for row_idx in range(start_row, len(df)):
            role_cell = df.iloc[row_idx, 0] if df.shape[1] > 0 else None
            count_cell = df.iloc[row_idx, 2] if df.shape[1] > 2 else None

            # Останавливаемся на пустой строке
            if role_cell is None or (isinstance(role_cell, float) and pd.isna(role_cell)):
                logger.info(f"Empty role at row {row_idx}, stopping staff parsing")
                break

            role_name = str(role_cell).strip()
            if not role_name:
                break

            staff_count = self._parse_int(count_cell)
            logger.info(f"Staff: role='{role_name}', count={staff_count}")
            
            records.append({
                'role_name': role_name,
                'staff_count': staff_count
            })

        return records

    def extract_expense_records(self, file_content: bytes) -> Dict[str, Any]:
        """Извлечение блока «Расходы» - горизонтальный формат"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=None, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading Excel for expense block: {e}")
            return {}

        if df.empty:
            return {}

        # Ищем заголовок "Расходы" в любой колонке
        expense_col = None
        start_row = None
        for row_idx in range(len(df)):
            for col_idx in range(df.shape[1]):
                cell = df.iloc[row_idx, col_idx]
                if isinstance(cell, str) and 'расход' in cell.strip().lower():
                    expense_col = col_idx
                    start_row = row_idx + 1
                    logger.info(f"Found 'Расходы' at row {row_idx}, col {col_idx}")
                    break
            if expense_col is not None:
                break

        if expense_col is None:
            logger.info("Expense block header not found")
            return {}

        records: List[Dict[str, Any]] = []
        calculated_total = Decimal('0.00')
        reported_total = None

        # Данные: колонка expense_col - статья, expense_col+2 - сумма (col+1 пустая)
        for row_idx in range(start_row, len(df)):
            item_cell = df.iloc[row_idx, expense_col] if df.shape[1] > expense_col else None
            
            if item_cell is None or (isinstance(item_cell, float) and pd.isna(item_cell)):
                break
            
            # Если в колонке статьи ЧИСЛО - это другой блок
            if isinstance(item_cell, (int, float)):
                logger.info(f"Row {row_idx}: col {expense_col} contains number {item_cell}, stopping expense parsing")
                break

            item_name = str(item_cell).strip()
            if not item_name:
                break

            # Ищем сумму справа (пропускаем пустые ячейки)
            amount = None
            for col_offset in range(1, 6):
                if df.shape[1] > expense_col + col_offset:
                    candidate = df.iloc[row_idx, expense_col + col_offset]
                    if candidate is not None and not (isinstance(candidate, float) and pd.isna(candidate)):
                        if isinstance(candidate, (int, float)):
                            amount = self._parse_decimal(candidate)
                            logger.info(f"Expense: item='{item_name}', amount={amount} (col offset {col_offset})")
                            break
                        elif isinstance(candidate, str):
                            # Текст вместо числа - может быть другой блок
                            if not candidate.replace('.', '').replace(',', '').replace('-', '').replace(' ', '').isdigit():
                                logger.info(f"Stopped at '{item_name}' - next column contains text '{candidate}'")
                                break
                            else:
                                amount = self._parse_decimal(candidate)
                                logger.info(f"Expense: item='{item_name}', amount={amount}")
                                break

            if amount is None:
                logger.info(f"No amount found for '{item_name}', skipping")
                continue

            # Проверяем на ИТОГО
            normalized = item_name.lower()
            if 'итого' in normalized:
                reported_total = amount
                records.append({
                    'expense_item': item_name,
                    'amount': amount,
                    'is_total': True
                })
                logger.info(f"Expense ИТОГО: {amount}")
                break

            calculated_total += amount
            records.append({
                'expense_item': item_name,
                'amount': amount,
                'is_total': False
            })

        if not records:
            return {}

        totals_match = True
        if reported_total is not None:
            if (reported_total - calculated_total).copy_abs() > Decimal('0.01'):
                totals_match = False
        else:
            reported_total = calculated_total

        return {
            'records': records,
            'calculated_total': calculated_total,
            'reported_total': reported_total,
            'totals_match': totals_match
        }

    def extract_cash_collection(self, file_content: bytes) -> Dict[str, Any]:
        """Извлечение блока «Инкассация» - горизонтальный формат"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=None, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading Excel for cash collection block: {e}")
            return {}

        if df.empty:
            return {}

        # Ищем заголовок "Инкассация" в любой колонке
        cash_col = None
        start_row = None
        for row_idx in range(len(df)):
            for col_idx in range(df.shape[1]):
                cell = df.iloc[row_idx, col_idx]
                if isinstance(cell, str) and 'инкассация' in cell.strip().lower():
                    cash_col = col_idx
                    start_row = row_idx + 1
                    logger.info(f"Found 'Инкассация' at row {row_idx}, col {col_idx}")
                    break
            if cash_col is not None:
                break

        if cash_col is None:
            logger.info("Cash collection block header not found")
            return {}

        # Пропускаем строку с заголовками (---, кол-во, курс, сумма)
        header_row = None
        for row_idx in range(start_row, min(start_row + 3, len(df))):
            cells = [df.iloc[row_idx, cash_col + i] if df.shape[1] > cash_col + i else None for i in range(4)]
            normalized = [str(cell).strip().lower() if cell is not None and not (isinstance(cell, float) and pd.isna(cell)) else '' for cell in cells]

            if any('кол' in cell for cell in normalized) or any('курс' in cell for cell in normalized):
                start_row = row_idx + 1
                logger.info(f"Found header row at {row_idx}, data starts at {start_row}")
                break

        records: List[Dict[str, Any]] = []
        calculated_total = Decimal('0.00')
        reported_total = None

        # Формат: cash_col - валюта, cash_col+1 - количество, cash_col+2 - курс, cash_col+3 - сумма
        for row_idx in range(start_row, len(df)):
            currency_cell = df.iloc[row_idx, cash_col] if df.shape[1] > cash_col else None
            quantity_cell = df.iloc[row_idx, cash_col + 1] if df.shape[1] > cash_col + 1 else None
            rate_cell = df.iloc[row_idx, cash_col + 2] if df.shape[1] > cash_col + 2 else None
            amount_cell = df.iloc[row_idx, cash_col + 3] if df.shape[1] > cash_col + 3 else None

            # Пропускаем пустые строки, ищем ИТОГО
            if currency_cell is None or (isinstance(currency_cell, float) and pd.isna(currency_cell)):
                # Проверяем следующие несколько строк на наличие ИТОГО
                found_total = False
                for offset in range(1, 8):
                    if row_idx + offset >= len(df):
                        break
                    next_cell = df.iloc[row_idx + offset, cash_col] if df.shape[1] > cash_col else None
                    if next_cell is not None and isinstance(next_cell, str) and 'итого' in next_cell.strip().lower():
                        # Нашли ИТОГО, продолжаем парсинг с этой строки
                        found_total = True
                        break
                if not found_total:
                    break
                else:
                    continue

            # Если в колонке валюты число - это другой блок
            if isinstance(currency_cell, (int, float)):
                logger.info(f"Row {row_idx}: col {cash_col} contains number, stopping cash parsing")
                break

            label = str(currency_cell).strip()
            if not label:
                continue

            normalized_label = label.lower()
            is_total = 'итого' in normalized_label

            quantity = None if is_total else self._parse_decimal(quantity_cell)
            rate = None if is_total else self._parse_decimal(rate_cell)
            amount = self._parse_decimal(amount_cell)

            # Вычисляем сумму, если не указана
            if not is_total and (amount is None or amount == Decimal('0.00')) and quantity is not None and rate is not None:
                amount = (quantity * rate).quantize(Decimal('0.01'))

            logger.info(f"Cash: currency='{label}', qty={quantity}, rate={rate}, amount={amount}, is_total={is_total}")

            records.append({
                'currency_label': label,
                'quantity': quantity,
                'exchange_rate': rate,
                'amount': amount,
                'is_total': is_total
            })

            if is_total:
                reported_total = amount
                break

            calculated_total += amount

        if not records:
            return {}

        if reported_total is None:
            reported_total = calculated_total

        totals_match = (reported_total - calculated_total).copy_abs() <= Decimal('0.01')

        return {
            'records': records,
            'calculated_total': calculated_total,
            'reported_total': reported_total,
            'totals_match': totals_match
        }

    def extract_staff_debts(self, file_content: bytes) -> Dict[str, Any]:
        """Извлечение блока «Долги по персоналу» - идет после инкассации БЕЗ заголовка"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=None, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading Excel for staff debts block: {e}")
            return {}

        if df.empty:
            return {}

        # Ищем ИТОГО инкассации, блок долгов идет сразу после него
        cash_itogo_row = None
        cash_col = None
        
        # Сначала ищем блок "Инкассация"
        for row_idx in range(len(df)):
            for col_idx in range(df.shape[1]):
                cell = df.iloc[row_idx, col_idx]
                if isinstance(cell, str) and 'инкассация' in cell.strip().lower():
                    # Нашли заголовок инкассации, ищем ИТОГО через 5-15 строк после него
                    for offset in range(5, 15):
                        if row_idx + offset >= len(df):
                            break
                        itogo_cell = df.iloc[row_idx + offset, col_idx] if df.shape[1] > col_idx else None
                        if itogo_cell and isinstance(itogo_cell, str) and 'итого' in itogo_cell.strip().lower():
                            # Проверяем, что справа есть сумма
                            amount_cell = df.iloc[row_idx + offset, col_idx + 3] if df.shape[1] > col_idx + 3 else None
                            if amount_cell is not None and isinstance(amount_cell, (int, float)):
                                cash_itogo_row = row_idx + offset
                                cash_col = col_idx
                                logger.info(f"Found cash ИТОГО at row {cash_itogo_row}, col {cash_col}, debts start after")
                                break
                    if cash_itogo_row is not None:
                        break
            if cash_itogo_row is not None:
                break

        if cash_itogo_row is None:
            logger.info("Staff debts block not found (no cash ИТОГО found)")
            return {}

        # Блок долгов начинается через 1-2 строки после ИТОГО инкассации
        start_row = cash_itogo_row + 2
        records = []
        calculated_total = Decimal('0.00')
        reported_total = None

        # Формат: cash_col - тип долга, cash_col+1 - сумма
        for row_idx in range(start_row, min(start_row + 10, len(df))):
            debt_type_cell = df.iloc[row_idx, cash_col] if df.shape[1] > cash_col else None
            amount_cell = df.iloc[row_idx, cash_col + 1] if df.shape[1] > cash_col + 1 else None

            # Останавливаемся на пустой строке
            if debt_type_cell is None or (isinstance(debt_type_cell, float) and pd.isna(debt_type_cell)):
                break

            # Если число в колонке типа - это другой блок
            if isinstance(debt_type_cell, (int, float)):
                break

            debt_type = str(debt_type_cell).strip()
            if not debt_type:
                break

            amount = self._parse_decimal(amount_cell)
            is_total = 'итого' in debt_type.lower()

            logger.info(f"Debt: type='{debt_type}', amount={amount}, is_total={is_total}")

            records.append({
                'debt_type': debt_type,
                'amount': amount,
                'is_total': is_total
            })

            if is_total:
                reported_total = amount
                break
            else:
                calculated_total += amount

        if not records:
            return {}

        if reported_total is None:
            reported_total = calculated_total

        totals_match = (reported_total - calculated_total).copy_abs() <= Decimal('0.01')

        return {
            'records': records,
            'calculated_total': calculated_total,
            'reported_total': reported_total,
            'totals_match': totals_match
        }

    def extract_notes_entries(self, file_content: bytes) -> Dict[str, List[Dict[str, Any]]]:
        """Извлечение блока «Примечание»"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=None, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading Excel for notes block: {e}")
            return {}

        if df.empty:
            return {}

        # Ищем заголовок "Примечания" в любой колонке
        start_row = None
        notes_col = None
        
        for row_idx in range(len(df)):
            for col_idx in range(df.shape[1]):
                cell = df.iloc[row_idx, col_idx]
                if isinstance(cell, str) and 'примечан' in cell.strip().lower():
                    start_row = row_idx + 1
                    notes_col = col_idx
                    logger.info(f"Found 'Примечания' at row {row_idx}, col {col_idx}")
                    break
            if start_row is not None:
                break

        if start_row is None or notes_col is None:
            logger.info("Notes block header not found")
            return {}

        column_headers_row = None
        for row_idx in range(start_row, len(df)):
            left_cell = df.iloc[row_idx, notes_col] if df.shape[1] > notes_col else None
            right_cell = df.iloc[row_idx, notes_col + 1] if df.shape[1] > notes_col + 1 else None

            if left_cell is None and right_cell is None:
                continue

            left_text = str(left_cell).strip().lower() if left_cell is not None else ''
            right_text = str(right_cell).strip().lower() if right_cell is not None else ''

            if 'долг' in left_text or 'долг' in right_text:
                column_headers_row = row_idx
                start_row = row_idx + 1
                logger.info(f"Found debt headers at row {row_idx}, data starts at {start_row}")
                break
            else:
                column_headers_row = row_idx
                start_row = row_idx
                break

        without_cash: List[Dict[str, Any]] = []
        with_cash: List[Dict[str, Any]] = []
        extra_notes: List[str] = []

        left_done = False
        right_done = False

        for row_idx in range(start_row, len(df)):
            left_cell = df.iloc[row_idx, notes_col] if df.shape[1] > notes_col else None
            right_cell = df.iloc[row_idx, notes_col + 1] if df.shape[1] > notes_col + 1 else None

            if left_cell is None and right_cell is None:
                continue

            left_text = str(left_cell).strip() if left_cell is not None and not (isinstance(left_cell, float) and pd.isna(left_cell)) else ''
            right_text = str(right_cell).strip() if right_cell is not None and not (isinstance(right_cell, float) and pd.isna(right_cell)) else ''

            left_lower = left_text.lower()
            right_lower = right_text.lower()
            
            # Останавливаемся если встречаем слова "доход", "расход", "прибыль" - это итоговый баланс
            if any(word in left_lower or word in right_lower for word in ['доход', 'расход', 'прибыль']):
                logger.info(f"Found balance keywords at row {row_idx}, stopping notes parsing")
                break

            processed_left = False
            processed_right = False

            if left_text and not left_done:
                if left_lower.startswith('итого'):
                    amount = self._parse_decimal(left_text.split(':')[-1])
                    without_cash.append({
                        'category': 'безнал',
                        'entry_text': left_text,
                        'is_total': True,
                        'amount': amount
                    })
                    left_done = True
                    processed_left = True
                else:
                    without_cash.append({
                        'category': 'безнал',
                        'entry_text': left_text,
                        'is_total': False
                    })
                    processed_left = True

            if right_text and not right_done:
                if right_lower.startswith('итого'):
                    amount = self._parse_decimal(right_text.split(':')[-1])
                    with_cash.append({
                        'category': 'нал',
                        'entry_text': right_text,
                        'is_total': True,
                        'amount': amount
                    })
                    right_done = True
                    processed_right = True
                else:
                    with_cash.append({
                        'category': 'нал',
                        'entry_text': right_text,
                        'is_total': False
                    })
                    processed_right = True

            if left_done and right_done and not (processed_left or processed_right):
                combined = " ".join(part for part in [left_text, right_text] if part).strip()
                if combined:
                    extra_notes.append(combined)

            elif not processed_left and left_text:
                extra_notes.append(left_text)

            elif not processed_right and right_text:
                extra_notes.append(right_text)

        return {
            'безнал': without_cash,
            'нал': with_cash,
            'extra': extra_notes
        }

    def extract_totals_summary(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Извлечение блока «Итоговый баланс» - горизонтальный формат"""
        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=None, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading Excel for totals summary block: {e}")
            return []

        if df.empty:
            return []

        # Ищем строку с заголовками "Доход", "Расход", "Чистая прибыль"
        balance_col = None
        start_row = None
        
        for row_idx in range(len(df)):
            for col_idx in range(df.shape[1]):
                cell = df.iloc[row_idx, col_idx]
                if isinstance(cell, str) and 'доход' in cell.strip().lower():
                    # Проверяем, что справа есть "Расход"
                    next_cell = df.iloc[row_idx, col_idx + 1] if df.shape[1] > col_idx + 1 else None
                    if next_cell and isinstance(next_cell, str) and 'расход' in next_cell.strip().lower():
                        balance_col = col_idx - 1  # Колонка с типом оплаты (левее "Дохода")
                        start_row = row_idx + 1
                        logger.info(f"Found totals header at row {row_idx}, col {col_idx}, data starts at {start_row}")
                        break
            if start_row is not None:
                break

        if start_row is None or balance_col is None:
            logger.info("Totals summary block header not found")
            return []

        expected_types = ['наличные', 'б/н', 'итого']
        records: List[Dict[str, Any]] = []

        # Формат: balance_col - тип оплаты, balance_col+1 - доход, balance_col+2 - расход, balance_col+3 - чистая прибыль
        for row_idx in range(start_row, min(start_row + 5, len(df))):
            type_cell = df.iloc[row_idx, balance_col] if df.shape[1] > balance_col else None
            income_cell = df.iloc[row_idx, balance_col + 1] if df.shape[1] > balance_col + 1 else None
            expense_cell = df.iloc[row_idx, balance_col + 2] if df.shape[1] > balance_col + 2 else None
            net_cell = df.iloc[row_idx, balance_col + 3] if df.shape[1] > balance_col + 3 else None

            if type_cell is None or (isinstance(type_cell, float) and pd.isna(type_cell)):
                break

            payment_type = str(type_cell).strip()
            lower_type = payment_type.lower()
            
            # Проверяем только если это ожидаемый тип
            if lower_type not in expected_types:
                break

            income = self._parse_decimal(income_cell)
            expense = self._parse_decimal(expense_cell)
            net = self._parse_decimal(net_cell)

            logger.info(f"Totals: type='{payment_type}', income={income}, expense={expense}, net={net}")

            records.append({
                'payment_type': payment_type,
                'income_amount': income,
                'expense_amount': expense,
                'net_profit': net
            })

        return records

    def export_off_shift_expenses_to_excel(self, expenses: List[Dict[str, Any]], club_name: str, start_date, end_date) -> bytes:
        """Экспорт расходов вне смены в Excel"""
        try:
            from datetime import date
            from openpyxl.styles import Font, Alignment
            from decimal import Decimal
            
            # Подготавливаем данные для DataFrame
            data = []
            for exp in expenses:
                data.append({
                    'Статья расхода': exp.get('expense_item', ''),
                    'Сумма': float(Decimal(str(exp.get('amount', 0))))
                })
            
            df = pd.DataFrame(data)
            
            # Добавляем итоговую строку
            total_amount = sum(Decimal(str(exp.get('amount', 0))) for exp in expenses)
            total_row = pd.DataFrame({
                'Статья расхода': ['ИТОГО'],
                'Сумма': [float(total_amount)]
            })
            df = pd.concat([df, total_row], ignore_index=True)
            
            # Создание Excel файла в памяти
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Записываем данные, начиная со строки 3 (оставляем место для заголовка)
                df.to_excel(writer, index=False, sheet_name='Расходы', startrow=2)
                
                # Получаем worksheet для форматирования
                worksheet = writer.sheets['Расходы']
                
                # Форматируем период
                start_str = start_date.strftime("%d.%m.%Y") if isinstance(start_date, date) else str(start_date)
                if start_date == end_date:
                    period_text = start_str
                else:
                    end_str = end_date.strftime("%d.%m.%Y") if isinstance(end_date, date) else str(end_date)
                    period_text = f"{start_str} - {end_str}"
                
                # Добавляем заголовок
                worksheet['A1'] = f'Дата (период): {period_text}'
                worksheet['A1'].font = Font(bold=True, size=12)
                
                worksheet['B1'] = f'Клуб: {club_name}'
                worksheet['B1'].font = Font(bold=True, size=12)
                
                # Делаем итоговую строку жирной
                bold_font = Font(bold=True, size=11)
                last_row = worksheet.max_row
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=last_row, column=col_idx)
                    cell.font = bold_font
                    if col_idx == 1:  # Колонка "Статья расхода"
                        cell.value = 'ИТОГО'
                    elif col_idx == 2:  # Колонка "Сумма"
                        cell.value = float(total_amount)
                
                # Выравнивание заголовков
                for col_idx in range(1, worksheet.max_column + 1):
                    header_cell = worksheet.cell(row=3, column=col_idx)
                    header_cell.font = Font(bold=True)
                    header_cell.alignment = Alignment(horizontal='center')
            
            output.seek(0)
            return output.getvalue()
        
        except Exception as e:
            logger.error(f"Error exporting off-shift expenses to Excel: {e}")
            raise ValueError(f"Не удалось экспортировать расходы: {str(e)}")


